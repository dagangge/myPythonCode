from PIL import Image
import openpyxl
from openpyxl.styles import PatternFill

im = Image.open('Pic2Excel\image.png')
width = im.size[0]
height = im.size[1]
im = im.convert('RGB')
array = []
for x in range(width):
    for y in range(height):
        r, g, b = im.getpixel((x,y))
        R = hex(r)[2:].ljust(2,'0')
        G = hex(g)[2:].ljust(2,'0')
        B = hex(b)[2:].ljust(2,'0')
        RGB = R+G+B
        if len(RGB)==6:
            array.append(RGB)
        else:
            print(RGB)
        

workbook = openpyxl.Workbook()
sheet = workbook.create_sheet(index=0, title="Report")

green_fill = PatternFill(bgColor="AACF91", fill_type="solid")
sheet.cell(row=1, column=2).fill = green_fill

for a in range(width):
    for b  in range(height):
        NumOfColor = a*height + b
        aaa = array[NumOfColor]
        Style_fill = PatternFill(fgColor= str(aaa), fill_type="solid")
        sheet.cell(row = b+1, column = a+1).fill = Style_fill
        

workbook.save('test.xlsx')
